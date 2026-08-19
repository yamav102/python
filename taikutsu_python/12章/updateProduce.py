#! python 3
# P315
# スプレッドシートを更新する
# updateProduce.py

# PRODUCE,	COST PER POUND,	POUNDS SOLD,	TOTAL
# 品種、1ポンド当たりの価格、売上ポンド数、総収入: round(価格/￡ * 売上￡数, 2)
# garlic, celery, lemon の「1ポンド当たりの価格」を変更する
# ファイルを別名で保存する

# 更新後の価格/￡
# Celery 1.19
# Garlic 3.07
# Lemon  1.27

import openpyxl
from pathlib import Path
import jaconv

# from openpyxl import Workbook
orginxl:str = 'produceSales.xlsx'
wb = openpyxl.load_workbook(Path(__file__).parent / orginxl)
sheet = wb['Sheet']

# 農産物の種類と、更新する価格
PRICE_UPDATES:dict = {'Garlic': 3.77,
                      'Celery': 1.99,
                      'Lemon': 1.77}

# 行をループして価格を更新する
for row_num in range(2, sheet.max_row + 1): # 先頭行をスキップ
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
updxl:str = 'updatedProduceSasles.xlsx'
savepath = Path(__file__).parent / updxl 
print(savepath)
if (
    not savepath.exists()
    # or input('上書きしますか？(y/n):').strip().lower() in ('y', 'yes', 'はい') # 'ｙ'を拾えない。
    or jaconv.z2h(input('上書きしますか？(y/n):').strip().lower(), ascii=True) in ('y', 'yes', 'はい')
   ):
    try:    
        wb.save(savepath)
        print('保存しました。')
    except PermissionError as e:
        print(f'Err:ファイルが開いています。\n⇒{savepath} \n⇒{e}')
    except Exception as e:
        print(f'Err:予期せぬエラー。\n⇒{savepath} \n⇒{e}')
    
else:
    print('保存をキャンセルしました。')



