#! python3
# P323 
# _12_9_2.python3
# load_workbook() では、数式をとるか、値をとるか、どちらかしか出来ない

import openpyxl
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet

fomulaxlpath = Path(__file__).parent / 'writeFormula.xlsx'
# data_only=False で数式を取得できる
# wb_formula = openpyxl.load_workbook(fomulaxlpath, data_only=False)
wb_formula = openpyxl.load_workbook(fomulaxlpath)
sheet = wb_formula.active
assert isinstance(sheet, Worksheet)
print(sheet['a3'].value) # =sum(a1:a2)

# wb_formula.close()

wb_data_only = openpyxl.load_workbook(fomulaxlpath, data_only=True)
sheet = wb_data_only.active
assert isinstance(sheet, Worksheet)
# openpyxl は数式を計算ので None が返る。一度エクセルで開いて、
# エクセルに計算させて保存して閉じた状態で実行すると、500が返る
print(sheet['a3'].value) # None 
