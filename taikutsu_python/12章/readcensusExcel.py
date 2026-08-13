# python3
# readcensusExcel.py
# 退屈な事はPythonにやらせよう P307 12.4
# censuspopdata.xlsx['Population by Census Tract']シート.Census Tract[国勢調査区]
# シート名：Population by Census Tract
# CensusTract:集計区ID（レコード数:72,864）、state:州、county:郡、POP2010:人口のリスト。異なる州で、同じ名前の郡がある場合があるので、州＋郡をキーで数えると 3,143件になる。
# census2010.py の作成を目的とする。census2010.py が出来てしまえば、本処理は不要で、census2010.pyをimport して、census2010.all_data[州名][郡名]
# P309
import openpyxl
import pprint # pprintモジュールをインポート
# pprint.pprint(type(pprint)) # <class 'module'>
# from pprint import pprint # pprintモジュールから pprint 関数をインポート
from pathlib import Path # pathlibモジュールから Path クラスをインポート
print('ワークブックを開いています...')
filename = 'censuspopdata.xlsx'
sheetname = 'Population by Census Tract'
wb = openpyxl.load_workbook(Path(__file__).parent / filename)
# with load_workbook(Path(__file__).parent / filename) as wb:とは書けない。
# print(dir(wb))
# print(help(wb))
# print(wb.path) # /xl/workbook.xml ディレクトリ上の path が返るわけではない。.xlsx を zip展開した中身の workbook.xmlの相対パスが返る。
# print(wb.name) # error　ブック名を返すプロパティは用意されていない。
sheet = wb[sheetname]
# print(sheet['A1'].value) # A1セルの値「CensusTract」
# print(sheet.title) # 'Population by Census Tract'

county_data = {} # 空の dict,空の集合の定義は s = set()、s = {1,2,3}

# TODO: conty_data に郡の人口と地域数を格納する
# P310
print('行を読み込んでいます...')
for row in range(2, sheet.max_row + 1):
    # スプレッドシートの1行に、ひとつの人口調査標準地域のデータがある
    state = sheet['B' + str(row)].value
    county = sheet['c' + str(row)].value
    pop = sheet['d' + str(row)].value

    # この州のキーを指定
    ## setdefault は、キーが存在すれば何もしないので、ループの中で何度実行されても問題は起きない。
    county_data.setdefault(state, {})
    # この州の{郡:{'tracts':初期値:0, 'pop':初期値:0}}を指定
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})
    # 各行の人口調査標準地域数を tracts として取得している。
    county_data[state][county]['tracts'] += 1
    # この人口調査標準地域の人口だけ郡の人口を増やす
    county_data[state][county]['pop'] += int(pop)

# wb.close() # read_only=True時は必須。それ以外は不要で無意味。
del sheet # 変数を手放すだけで、メモリが開放されるわけではないが、ガベージコレクションの対象になる。
del wb

# TODO: 新しいテキストファイルを開き county_dataの内容を書き込む

# P312
print('結果を書き込み中...')
with open(Path(__file__).parent / 'census2010.py', 'w') as result_file:
    result_file.write('all_data = ' + pprint.pformat(county_data)) # pprint.pformat は pprint の結果を文字列として返す
    # pprint は key でソートして出力がdefault。'pop' →'tracts'の順での出力になる。
# all_data = {'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
#         'Aleutians West': {'pop': 5561, 'tracts': 2},
#         'Anchorage': {'pop': 291826, 'tracts': 55},
#         'Bethel': {'pop': 17013, 'tracts': 3},    
print('完了')

